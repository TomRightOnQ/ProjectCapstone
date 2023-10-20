import csv
import os
from openpyxl import load_workbook

'''Parse csv or xlsx file to consts'''

def wrap_quotes_if_string(value, data_type):
    if isinstance(value, bool):
        value = str(value).lower()
    if not isinstance(value, str):
        value = str(value)
    if data_type == 'string':
        return f'"{value}"'
    elif data_type == 'float':
        return f"{value}f"
    elif data_type == 'int[]':
        return f"new int[]{{{value}}}"
    elif data_type == 'float[]':
        return f"new float[]{{{value}}}"
    elif data_type == 'string[]':
        return f"new string[]{{{value}}}"
    elif data_type == 'Vector3':
        return f"new Vector3({value})"
    else:
        return value

def convert_to_cs(dataHeader, dataType, lines, output_filename):
    className = os.path.basename(output_filename).replace('Data.cs', 'Data')
    structName = className + 'Struct'
    
    with open(output_filename, 'w') as outfile:
        outfile.write('using System.Collections.Generic;\n\n')
        outfile.write('using UnityEngine;\n\n')
        outfile.write(f'public static class {className}\n')
        outfile.write('{\n')
        # Define MyData class
        outfile.write(f'    public class {structName}\n')
        outfile.write('    {\n')
        for h, t in zip(dataHeader, dataType):
            outfile.write(f'        public {t} {h};\n')
        outfile.write('\n')
        
        # Add constructor
        constructor_args = ', '.join([f"{t} {h}" for h, t in zip(dataHeader, dataType)])
        constructor_body = '\n'.join([f"            this.{h} = {h};" for h in dataHeader])
        outfile.write(f'        public {structName}({constructor_args})\n')
        outfile.write('        {\n')
        outfile.write(constructor_body)
        outfile.write('\n        }\n')
        outfile.write('    }\n')
        
        # Create dictionary
        outfile.write(f'    public static Dictionary<int, {structName}> data = new Dictionary<int, {structName}>\n')
        outfile.write('    {\n')
        
        for line in lines:
            values = [wrap_quotes_if_string(value, t) for value, t in zip(line, dataType)]
            id_value = values[0]
            other_values = ', '.join(values[1:])
            outfile.write(f'        {{{id_value}, new {structName}({id_value}, {other_values})}},\n')
        outfile.write('    };\n')

        # Add a getter
        outfile.write(f"""
    public static {structName} GetData(int id)
    {{
        if (data.TryGetValue(id, out {structName} result))
        {{
            return result;
        }}
        else
        {{
            return null;
        }}
    }}
""")
        outfile.write('}\n')

def convert_csv_to_cs(csv_filename, output_filename):
    with open(csv_filename, 'r') as f:
        reader = csv.reader(f)
        lines = list(reader)
        dataHeader = lines[0]
        dataType = lines[1]
        lines = lines[2:]
        
        convert_to_cs(dataHeader, dataType, lines, output_filename)

def convert_xlsx_to_cs(xlsx_filename, output_filename):
    workbook = load_workbook(xlsx_filename)
    worksheet = workbook.active

    dataHeader = [cell.value for cell in worksheet[1]]
    dataType = [cell.value for cell in worksheet[2]]
    lines = [[cell.value for cell in row] for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row)]
    
    convert_to_cs(dataHeader, dataType, lines, output_filename)

def is_temporary_excel_file(filename):
    return filename.startswith("~$")

root_dir = os.path.join(os.path.dirname(__file__), '..', '..')

data_folder = os.path.join(root_dir, 'Designs', 'General')
output_folder = os.path.join(root_dir, 'Scripts', 'Contents', 'General')

# Ensure output folder exists
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Loop through each file in the directory
for filename in os.listdir(data_folder):
    input_filename = os.path.join(data_folder, filename)
    output_filename = os.path.join(output_folder, filename.split('.')[0] + 'Data.cs')
    if is_temporary_excel_file(filename):
        continue
    if filename.endswith('.csv'):
        convert_csv_to_cs(input_filename, output_filename)
    elif filename.endswith('.xlsx'):
        convert_xlsx_to_cs(input_filename, output_filename)

# Additional code here if needed
print("All files have been converted.")
